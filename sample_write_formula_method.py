import openpyxl
import pprint

book = openpyxl.Workbook()

"""
Excelのセルに記入するデータ
"""
max_row = 2
max_col = 2
itemlist = ["apple", "orange"]
owner_list = ["hero", "tom"]
item_old_price = ["100", "200"]
item_new_price = ["200", "200"]

"""
シート「前版の内容」にデータを入力
"""
old_value_sheet = book.create_sheet('前版の内容')
old_value_sheet["A1"] = "前版の内容"

old_value_sheet["A3"] = "検索用文字列"
old_value_sheet["B3"] = "項目"
old_value_sheet["C3"] = "生産者"
old_value_sheet["D3"] = "データ"

# 項目を入力
for r in range(max_row):
  old_value_sheet.cell(row=r+4, column=2).value = itemlist[r]

# 生産者を入力
for r in range(max_row):
  old_value_sheet.cell(row=r+4, column=3).value = owner_list[r]

# データを入力
for r in range(max_row):
  old_value_sheet.cell(row=r+4, column=4).value = item_old_price[r]

# 検索用文字列を入力
for r in range(max_row):
  search_str = str(old_value_sheet.cell(row=r+4, column=2).value) + str(old_value_sheet.cell(row=r+4, column=3).value)
  old_value_sheet.cell(row=r+4, column=1).value = search_str

"""
シート「次版の内容」にデータを入力
"""
new_value_sheet = book.create_sheet('次版の内容')
new_value_sheet["A1"] = "次版の内容"

new_value_sheet["A3"] = "検索用文字列"
new_value_sheet["B3"] = "項目"
new_value_sheet["C3"] = "生産者"
new_value_sheet["D3"] = "データ"

# 項目を入力
for r in range(max_row):
  new_value_sheet.cell(row=r+4, column=2).value = itemlist[r]

# 生産者を入力
for r in range(max_row):
  new_value_sheet.cell(row=r+4, column=3).value = owner_list[r]

# データを入力
for r in range(max_row):
  new_value_sheet.cell(row=r+4, column=4).value = item_new_price[r]

# 検索用文字列を入力
for r in range(max_row):
  search_str = str(new_value_sheet.cell(row=r+4, column=2).value) + str(new_value_sheet.cell(row=r+4, column=3).value)
  new_value_sheet.cell(row=r+4, column=1).value = search_str

"""
シート「環境変数比較」にデータを入力
"""
env_compare_sheet = book.create_sheet('環境変数比較')
env_compare_sheet['A1'] = "比較"

env_compare_sheet['A3'] = "項目"
env_compare_sheet['B3'] = "生産者"
env_compare_sheet['C3'] = "前版のデータ"
env_compare_sheet['D3'] = "次版のデータ"
env_compare_sheet['E3'] = "比較結果"


# 項目を入力
for r in range(max_row):
  env_compare_sheet.cell(row=r+4, column=1).value = itemlist[r]

# 生産者を入力
for r in range(max_row):
  env_compare_sheet.cell(row=r+4, column=2).value = owner_list[r]


# 前版のデータを入力(関数の書き込み)
for r in range(max_row):
  search_word = str(env_compare_sheet.cell(row=r+4, column=1).value) + str(env_compare_sheet.cell(row=r+4, column=2).value)
  siki = "=VLOOKUP(\"" + str(search_word) + "\",前版の内容!A4:D5,4,0)" # 関数の作成
  env_compare_sheet.cell(row=r+4, column=3).value = siki

# 次版のデータを入力(関数の書き込み)
for r in range(max_row):
  search_word = str(env_compare_sheet.cell(row=r+4, column=1).value) + str(env_compare_sheet.cell(row=r+4, column=2).value)
  siki = "=VLOOKUP(\"" + str(search_word) + "\",次版の内容!A4:D5,4,0)" # 関数の作成
  env_compare_sheet.cell(row=r+4, column=4).value = siki

# 比較結果を入力(数式の書き込み)
for r in range(max_row):
  siki = "=IF(C" + str(r+4) + "=D" + str(r+4) + ",\"ok\",\"NG\")" # 数式の作成
  env_compare_sheet.cell(row=r+4, column=5).value = siki


book.save('sample.xlsx') # Excelファイルを保存
